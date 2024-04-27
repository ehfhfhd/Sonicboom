import glob
import json
import pandas as pd
import datetime
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# 현재 디렉토리에서 "result"로 시작하는 모든 JSON 파일의 경로 찾기
json_files = glob.glob('result*.json')

data = {} # 진단 데이터(json)
df = {}   # 진단 데이터프레임
df_info = {}
df_results = {}
num = 0   # 진단 서버 개수 

# 각 결과를 데이터프레임으로 저장하기
for json_file in json_files:
    with open(json_file, 'r', encoding='utf-8') as f:
        data[num] = json.load(f)
        df_info[num] = pd.json_normalize(data[num]['Server_Info'])
        df_results[num] = pd.json_normalize(data[num]['Check_Results'])

        # info 데이터 프레임에 진단 서버 번호 칼럼 넣기
        df_info[num].insert(0, 'No', num+1)

        # 데이터프레임 칼럼 이름 변경
        df_info[num] = df_info[num].rename(columns={"SW_TYPE":"타입", "SW_NM":"운영체제", "SW_INFO":"운영체제 정보",
                                                    "HOST_NM":"호스트명", "DATE":"진단 일", "TIME":"진단 시간", "IP_ADDRESS":"IP 주소"})
        df_results[num] = df_results[num].rename(columns={"Category":"항목 분류", "Item":"항목 코드", "Sub_Category":"진단 항목",
                                                          "Importance":"위험도", "status": "진단 결과", "Description":"판단 기준", "details":"현황", "solutions":"대응 방안"})

        # info 값을 results 줄 개수 만큼 복사 뒤 칼럼 합치기
        dftmp = df_info[num]
        for i in range(0, df_results[num]['항목 코드'].count() - 1):
            dftmp = pd.concat([dftmp, df_info[num]], ignore_index = True)
        df[num] = pd.concat([dftmp, df_results[num]], axis=1)

        # 진단 서버 카운트
        num = num+1


# 모든 서버의 진단 데이터프레임 합치기
df_all = df[0]
df_info_all = df_info[0]
for i in range(1, num):
    df_all = pd.concat([df_all, df[i]])
    df_info_all = pd.concat([df_info_all, df_info[i]])
df_all = df_all[["No", "운영체제", "호스트명", "진단 일", "IP 주소",
                 "항목 코드", "진단 항목", "위험도", "진단 결과", "판단 기준", "현황", "대응 방안"]]

# 엑셀 파일에서 보기 좋게 엔터키로 구분
df_all['현황'] = df_all['현황'].apply(lambda x: '\n'.join(x))
df_all['대응 방안'] = df_all['대응 방안'].apply(lambda x: '\n'.join(x) if isinstance(x, list) else '') # 리스트가 있는 경우에만



# 현재 날짜 가져오기
current_date = datetime.datetime.now().date()
date_str = current_date.strftime("%Y%m%d") # 문자열로 변환
exel_name = f'result_{date_str}.xlsx'

# 엑셀 템플릿 복사한 뒤 이름을 현재 날짜와 연관하여 저장하기
shutil.copy('tmp.xlsx', exel_name)

# 데이터 프레임 엑셀에 저장하기
with pd.ExcelWriter('df.xlsx', engine='xlsxwriter') as writer:
    df_info_all.to_excel(writer, sheet_name = '1. 진단대상', startrow = 2, header = True, index = False)
    df_all.to_excel(writer, sheet_name = '5. 진단결과 상세', startrow = 2, header = True, index = False)

# 엑셀 파일 로드
workbook = load_workbook(exel_name)
workbookdf = load_workbook('df.xlsx')

# 시트 복제
dfsheet_1 = workbookdf['1. 진단대상']
dfsheet_5 = workbookdf['5. 진단결과 상세']
sheet_1 = workbook.create_sheet('1. 진단대상')
sheet_5 = workbook.create_sheet('5. 진단결과 상세')
for row in dfsheet_1.iter_rows():
    for cell in row:
        sheet_1[cell.coordinate].value = cell.value
for row in dfsheet_5.iter_rows():
    for cell in row:
        sheet_5[cell.coordinate].value = cell.value

# 시트 위치 바꾸기
workbook.move_sheet('1. 진단대상', -3)


# 시트 1번 수정
# 제목
sheet_1.cell(1,1, '1. 진단대상')
sheet_1.cell(1,1).font = Font(bold=True, size=14)
# 결과 칼럼의 너비 조절
sheet_1.column_dimensions['A'].width = 5
sheet_1.column_dimensions['D'].width = 35
sheet_1.column_dimensions['E'].width = 20
sheet_1.column_dimensions['H'].width = 15
sheet_1.column_dimensions['I'].width = 25
# 모든 셀에 대해 수직 및 수평 정렬 설정 및 테두리 추가
for row in sheet_1.iter_rows(min_row=3, max_row=sheet_1.max_row, min_col=1, max_col=sheet_1.max_column):
    for cell in row:
        cell.alignment = Alignment(vertical='center', horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
# 칼럼 설정
for col_row in sheet_1.iter_rows(min_row=3, max_row=3, min_col=1, max_col=sheet_1.max_column):
    for col_cell in col_row:
        col_cell.fill = PatternFill(start_color='C6DAF1', end_color='C6DAF1', fill_type='solid') # 하늘색 바탕
        col_cell.font = Font(bold=True, size=12)


# 시트 5번 수정
# 제목
sheet_5.cell(1,1, '5. 진단결과 상세')
sheet_5.cell(1,1).font = Font(bold=True, size=14)
# 결과 칼럼의 너비 조절
sheet_5.column_dimensions['A'].width = 5
sheet_5.column_dimensions['C'].width = 20
sheet_5.column_dimensions['E'].width = 15
sheet_5.column_dimensions['F'].width = 10
sheet_5.column_dimensions['G'].width = 55
sheet_5.column_dimensions['I'].width = 10
sheet_5.column_dimensions['J'].width = 100
sheet_5.column_dimensions['K'].width = 100
sheet_5.column_dimensions['L'].width = 100
# 모든 셀에 대해 수직 및 수평 정렬 설정 및 테두리 추가
for row in sheet_5.iter_rows(min_row=3, max_row=sheet_5.max_row, min_col=1, max_col=sheet_5.max_column):
    for cell in row:
        cell.alignment = Alignment(vertical='center', horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
# 칼럼 설정
for col_row in sheet_5.iter_rows(min_row=3, max_row=3, min_col=1, max_col=sheet_5.max_column):
    for col_cell in col_row:
        col_cell.fill = PatternFill(start_color='C6DAF1', end_color='C6DAF1', fill_type='solid') # 하늘색 바탕
        col_cell.font = Font(bold=True, size=12)
# 칼럼에 필터 설정
sheet_5.auto_filter.ref = "A3:I3"
# 긴 부분은 개행으로 보기 좋게 설정
for row5 in sheet_5.iter_rows(min_row=4, max_row=sheet_5.max_row, min_col=7, max_col=7):
    for cell5 in row5:
        cell5.alignment = Alignment(vertical='center', horizontal='left')
for long_row in sheet_5.iter_rows(min_row=4, max_row=sheet_5.max_row, min_col=10,  max_col=12):
    for long_cell in long_row:
        long_cell.alignment = Alignment(vertical='center', horizontal='left', wrapText=True)
# 양호판단 결과 스타일 설정
for status_row in sheet_5.iter_rows(min_row=4, max_row=sheet_5.max_row, min_col=9, max_col=9):
    for status_cell in status_row:
        if status_cell.value == "[취약]":
            status_cell.font = Font(color="FF0000") # 빨간 글씨
        elif status_cell.value == "[인터뷰]":
            status_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # 노란색 바탕


# 시트 4번 수정
sheet_4 = workbook['4. 진단결과 요약']
# 유저 작성
sht4_user = df_info_all[["No", "진단 일", "IP 주소", "운영체제", "호스트명"]]
sht4_user_array = sht4_user.values
for n in range(0,num):
    for idx in range(0,5):
        sheet_4.cell(idx+3,n+5, sht4_user_array[n][idx])
        sheet_4.cell(idx+3,n+5).alignment = Alignment(vertical='center', horizontal='center')
        sheet_4.cell(idx+3,n+5).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                top=Side(style='thin'), bottom=Side(style='thin'))
    sheet_4.cell(3,n+5).font = Font(color="0000FF") # 빨간 글씨
    sheet_4.cell(7,n+5).alignment = Alignment(vertical='top', horizontal='left', wrapText=True)
    # 유저 배경
    for user_row in sheet_4.iter_rows(min_row=3, max_row=7, min_col=n+5, max_col=n+5):
        for user_cell in user_row:
            user_cell.fill = PatternFill(start_color='CCC0DA', end_color='CCC0DA', fill_type="solid") # 보라색 바탕
    # 진단 결과 칼럼
    sheet_4.cell(8,n+5, '진단 결과')
    sheet_4.cell(8,n+5).font = Font(bold=True)
    # 진단 결과 작성
    for idx in range(0,72):
        sheet_4.cell(idx+9,n+5, df[n]['진단 결과'][idx])
    # 양호판단 결과 스타일 설정
    for status_row in sheet_4.iter_rows(min_row=9, max_row=80, min_col=n+5, max_col=n+5):
        for status_cell in status_row:
            if status_cell.value == "[취약]":
                status_cell.font = Font(color="FF0000") # 빨간 글씨
            elif status_cell.value == "[인터뷰]":
                status_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # 노란색 바탕
    # 진단 결과 통계
    ascii_4 = chr(69+n) # 아스키코드화
    sheet_4.cell(81,n+5, f'=COUNTIF({ascii_4}9:{ascii_4}80, "[양호]")') # 양호 개수
    sheet_4.cell(82,n+5, f'=COUNTIF({ascii_4}9:{ascii_4}80, "[취약]")') # 취약 개수
    sheet_4.cell(83,n+5, f'=72-{ascii_4}81-{ascii_4}82') # N/A 개수
    sheet_4.cell(81,n+5).fill = PatternFill(start_color='C6DAF1', end_color='C6DAF1', fill_type='solid') # 하늘색 바탕
    sheet_4.cell(82,n+5).fill = PatternFill(start_color='C6DAF1', end_color='C6DAF1', fill_type='solid')
    sheet_4.cell(83,n+5).fill = PatternFill(start_color='C6DAF1', end_color='C6DAF1', fill_type='solid')
    # 보안 수준
    sheet_4.cell(84,n+5,
                 f'=SUMIF({ascii_4}9:{ascii_4}80,"[양호]",\'3. 영역별 보안지수\'!A4:A75)/(SUM(\'3. 영역별 보안지수\'!A4:A75)-SUMIF({ascii_4}9:{ascii_4}80,"[인터뷰]",\'3. 영역별 보안지수\'!A4:A75)-SUMIF({ascii_4}9:{ascii_4}80,"[N/A]",\'3. 영역별 보안지수\'!A4:A75))')
    sheet_4.cell(84,n+5).number_format = '0.0%' # 표시 형식 변경
    sheet_4.cell(84,n+5).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # 노란색 바탕
    # 가운데 정렬 및 테두리 작성
    for row in sheet_4.iter_rows(min_row=8, max_row=84, min_col=n+5, max_col=n+5):
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))




# 시트 3번 수정
sheet_3 = workbook['3. 영역별 보안지수']
for row_3 in range(4,76):
    sheet_3.cell(row_3,3, f'=COUNTIF(\'4. 진단결과 요약\'!E{row_3+5}:{chr(68+num)}{row_3+5}, "[양호]")') # 영역 별 양호 수
    sheet_3.cell(row_3,4, f'=COUNTIF(\'4. 진단결과 요약\'!E{row_3+5}:{chr(68+num)}{row_3+5}, "[취약]")') # 영역 별 취약 수
    sheet_3.cell(row_3,5, f'=COUNTIF(\'4. 진단결과 요약\'!E{row_3+5}:{chr(68+num)}{row_3+5}, "[인터뷰]")+COUNTIF(\'4. 진단결과 요약\'!E{row_3+5}:{chr(68+num)}{row_3+5}, "[N/A]")') # 영역 별 N/A 수



# 표지 수정
sheet_0 = workbook['표지']
sheet_0.cell(18,1,current_date)
sheet_0.cell(18,1).number_format = 'yyyy"年" mm"月" dd"日"' # 표시 형식 변경


# 변경 사항 저장
workbook.save(exel_name)

