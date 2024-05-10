import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# JSON 파일 읽기
with open('result.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# 데이터프레임 생성
df = pd.json_normalize(data, 'Check_Results')

# 데이터프레임 칼럼 순서 및 이름 변경
df = df[["Category", "Item", "Importance", "status", "Sub_Category", "Description", "details", "solutions"]]
df = df.rename(columns={"Category":"항목분류", "Item":"코드", "Importance":"중요도", "status": "양호판단", "Sub_Category":"항목명",
                   "Description":"판단기준", "details":"현황", "solutions":"대응방안"})

# 엑셀 파일에서 보기 좋게 엔터키로 구분
df['항목명'] = df['항목명'].apply(lambda x: '[항목명]\n' + x + '\n')
df['판단기준'] = df['판단기준'].apply(lambda x: '\n[판단기준]\n' + x + '\n')
df['현황'] = df['현황'].apply(lambda x: '\n[현황]\n' + '\n'.join(x))
df['대응방안'] = df['대응방안'].apply(lambda x: '\n[대응방안]\n' + '\n'.join(x) if isinstance(x, list) else '') # 리스트가 있는 경우에만

# 결과 값 취합
df['결과'] = df['항목명'] + df['판단기준'] + df['현황'] + df['대응방안']
df = df[["항목분류", "코드", "중요도", "양호판단", "결과"]]

# 멀티인덱스 설정
df.set_index(["항목분류", "코드"], inplace=True)

# Excel 파일 생성 및 데이터프레임 저장
with pd.ExcelWriter('test2.xlsx', engine='xlsxwriter') as writer:
    df.to_excel(writer, sheet_name='Linux', startrow=1, header=True, index=True, index_label=["항목분류", "코드"])

    # 결과 칼럼의 너비 조절
    worksheet = writer.sheets['Linux']
    worksheet.set_column('A:A', 12)
    worksheet.set_column('E:E', 140)

# Excel 파일 열기
workbook = load_workbook('test2.xlsx')
sheet = workbook['Linux']

# 모든 셀에 대해 수직 및 수평 정렬 설정
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column - 1):
    for cell in row:
        cell.alignment = Alignment(vertical='center', horizontal='center')

# 변경 사항 저장
workbook.save('test2.xlsx')
